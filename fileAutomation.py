import openpyxl as xl
import generateCharts as gc
import helpFunctions as hf

from datetime import datetime

# function that gets file src and return a workbook object 'wb' from the loaded file
def load_workbook(filename):
    #Load
    return xl.load_workbook(filename)


def payment_method_by_gender(wb, sheet, exportFilename):
    cashF = 0
    cashM = 0
    cardF = 0
    cardM = 0
    ewalletF = 0
    ewalletM = 0

    paymentMethodTuple = ("Ewallet", "Cash", "Credit card")
    genderTuple = ("Male", "Female")

    # for lop that run on rows
    for row in range(2, sheet.max_row + 1):
        genderCell = sheet.cell(row, 5)
        paymentCell = sheet.cell(row, 13)

        if paymentCell.value == paymentMethodTuple[0]:
            if genderCell.value == genderTuple[1]:
                ewalletF += 1
            if genderCell.value == genderTuple[0]:
                ewalletM += 1
        if paymentCell.value == paymentMethodTuple[1]:
            if genderCell.value == genderTuple[1]:
                cashF += 1
            if genderCell.value == genderTuple[0]:
                cashM += 1
        if paymentCell.value == paymentMethodTuple[2]:
            if genderCell.value == genderTuple[1]:
                cardF += 1
            if genderCell.value == genderTuple[0]:
                cardM += 1

    maleTotal = cashM + cardM + ewalletM
    femaleTotal = cashF + cardF + ewalletF

    percentageCashF = hf.get_percentage(cashF, femaleTotal)
    percentageCashM = hf.get_percentage(cashM, maleTotal)

    percentageCardF = hf.get_percentage(cardF, femaleTotal)
    percentageCardM = hf.get_percentage(cardM, maleTotal)

    percentageWalletF = hf.get_percentage(ewalletF, femaleTotal)
    percentageWalletM = hf.get_percentage(ewalletM, maleTotal)

    #add new sheet
    sheet_data = (
        ("", paymentMethodTuple[1]+"%", paymentMethodTuple[2]+"%", paymentMethodTuple[0]+"%"),
        (genderTuple[0], percentageCashM, percentageCardM, percentageWalletM),
        (genderTuple[1], percentageCashF, percentageCardF, percentageWalletF)
    )
    new_sheet = hf.add_sheet_to_wb(wb, sheet_data, "PaymentMethodByGender")
    gc.payment_bar_chart(wb,new_sheet)
    wb.save(exportFilename)



def sales_by_category(wb, sheet, exportFilename):
    fashion = 0
    food = 0
    electronic = 0
    sport = 0
    home = 0
    health = 0
    categoriesTuple = ("Fashion accessories", "Food and beverages", "Electronic accessories", "Sports and travel",
                  "Home and lifestyle", "Health and beauty")

    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        categoryCell = sheet.cell(row,6)
        totalCell = sheet.cell(row,10)

        if categoryCell.value == categoriesTuple[0]:
            fashion += int(totalCell.value)
        if categoryCell.value == categoriesTuple[1]:
            food += int(totalCell.value)
        if categoryCell.value == categoriesTuple[2]:
            electronic += int(totalCell.value)
        if categoryCell.value == categoriesTuple[3]:
            sport += int(totalCell.value)
        if categoryCell.value == categoriesTuple[4]:
            home += int(totalCell.value)
        if categoryCell.value == categoriesTuple[5]:
            health += int(totalCell.value)


    #add new sheet
    sheet_data = (
        ("", "Total Sales(US Dollars)"),
        (categoriesTuple[0], electronic),
        (categoriesTuple[1], fashion),
        (categoriesTuple[2], food),
        (categoriesTuple[3], health),
        (categoriesTuple[4], home),
        (categoriesTuple[5], sport)
    )
    new_sheet = hf.add_sheet_to_wb(wb, sheet_data, "SalesByCategory")

    #call charts functions
    gc.category_bar_chart(wb,new_sheet)
    gc.category_pie_chart(wb,new_sheet)
    wb.save(exportFilename)


def sales_by_month(wb, sheet, exportFilename):
    #stores sales by month using dictionary
    data = hf.create_month_dict();

    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        totalCell = sheet.cell(row,10)
        dateCell = sheet.cell(row,11)

        #convert date from string to date object
        date_obj = hf.string_to_date(dateCell.value)

        if date_obj.month == 1:
            data["January"] = data["January"] + totalCell.value
        if date_obj.month == 2:
            data["February"] = data["February"] + totalCell.value
        if date_obj.month == 3:
            data["March"] = data["March"] + totalCell.value
        if date_obj.month == 4:
            data["April"] = data["April"] + totalCell.value
        if date_obj.month == 5:
            data["May"] = data["May"] + totalCell.value
        if date_obj.month == 6:
            data["June"] = data["June"] + totalCell.value
        if date_obj.month == 7:
            data["July"] = data["July"] + totalCell.value
        if date_obj.month == 8:
            data["August"] = data["August"] + totalCell.value
        if date_obj.month == 9:
            data["September"] = data["September"] + totalCell.value
        if date_obj.month == 10:
            data["October"] = data["October"] + totalCell.value
        if date_obj.month == 11:
            data["November"] = data["November"] + totalCell.value
        if date_obj.month == 12:
            data["December"] = data["December"] + totalCell.value

    #round data values
    data = hf.round_dict_values(data);

    #create new sheet

    sheet_data = (
        ('', 'Total Sales(US Dollar)'),
        ("January", data["January"]),
        ("February", data["February"]),
        ("March", data["March"]),
        ("April", data["April"]),
        ("May", data["May"]),
        ("June", data["June"]),
        ("July", data["July"]),
        ("August", data["August"]),
        ("September", data["September"]),
        ("October", data["October"]),
        ("November", data["November"]),
        ("December", data["December"])
    )
    new_sheet = hf.add_sheet_to_wb(wb, sheet_data, "SalesByMonth")

    # call line chart func
    gc.sales_line_chart(wb,new_sheet)

    wb.save(exportFilename)



def net_income(wb, sheet, exportFilename):
    taxes = 0
    total = 0

    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        tax_cell = sheet.cell(row,9)
        total_cell = sheet.cell(row, 10)

        taxes += int(tax_cell.value)
        total += int(total_cell.value)

    #create new sheet
    sheet_data = (
        ("", "Total Taxes", "Total Income", "Total NetIncome"),
        ("US Dollar",taxes,total,(total-taxes))
    )
    new_sheet = hf.add_sheet_to_wb(wb,  sheet_data, "NetIncome")

    wb.save(exportFilename)