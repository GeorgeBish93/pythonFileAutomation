from openpyxl.chart import BarChart,PieChart,Reference


def customers_type(wb, exportFilename):
    # access the sheet in the wb
    sheet = wb['Worksheet']
    member = 0
    normal = 0
    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        cell = sheet.cell(row,4)

        if cell.value == "Member":
            member += 1
        if cell.value == "Normal":
            normal += 1

    wb.create_sheet('CustomerType')
    new_sheet = wb['CustomerType']
    data = (
        ("Customer Type","Normal Customer", "Member Customer"),
        ("Total Customers",normal,member)
    )
    for i in data:
        new_sheet.append(i)

    values = Reference(new_sheet,
                       min_row=2,
                       max_row=2,
                       min_col=2,
                       max_col=3)

    chart = BarChart()
    chart.add_data(values,titles_from_data=True)
    chart.title = 'Customers By Membership'
    chart.x_axis.title = 'Customer Type'
    chart.y_axis.title = 'Number of Customers'
    chart.legend.position = 'r'
    new_sheet.add_chart(chart, 'f2')

    wb.save(exportFilename)


def customers_payment_method(wb, exportFilename):
    sheet = wb['Worksheet']
    cash = 0
    card = 0
    ewallet = 0

    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        cell = sheet.cell(row, 13)

        if cell.value == "Ewallet":
            ewallet += 1
        if cell.value == "Cash":
            cash += 1
        if cell.value == "Credit card":
            card += 1

    wb.create_sheet('PaymentMethod')
    new_sheet = wb['PaymentMethod']
    data = (
        ("", "Cash", "Credit Card", "Ewallet"),
        ("Total Customers", cash,card, ewallet )
    )
    for i in data:
        new_sheet.append(i)

    #create the chart and the references
    chart = PieChart()
    data = Reference(new_sheet, min_col=2, max_col=4, min_row=2, max_row=2)
    labels = Reference(new_sheet, min_col=2, max_col=4, min_row=1)

    #set the title, add the data to the chart and set the categories
    chart.title = 'Customer Payment Method Distribution'
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    #add the chart and save the workbook as a new file
    new_sheet.add_chart(chart, 'F5')
    wb.save(exportFilename)

def branch_customers(wb, exportFilename):
    sheet = wb['Worksheet']
    branchA = 0
    branchB = 0
    branchC = 0

    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        cell = sheet.cell(row,2)

        if cell.value == "A":
            branchA += 1
        if cell.value == "B":
            branchB += 1
        if cell.value == "C":
            branchC += 1

    wb.create_sheet('BranchCustomers')
    new_sheet = wb['BranchCustomers']
    data = (
        ("Branch","Total Customers"),
        ("A",branchA),
        ("B", branchB),
        ("C", branchC)
    )
    for i in data:
        new_sheet.append(i)

    wb.save(exportFilename)

def category_sales(wb, exportFilename):
    sheet = wb['Worksheet']
    fashion = 0
    food = 0
    electronic = 0
    sport = 0
    home = 0

    #for lop that run on rows
    for row in range(2,sheet.max_row +1):
        cell = sheet.cell(row,6)

        if cell.value == "Fashion accessories":
            fashion += 1
        if cell.value == "Food and beverages":
            food += 1
        if cell.value == "Electronic accessories":
            electronic += 1
        if cell.value == "Sports and travel":
            sport += 1
        if cell.value == "Home and lifestyle":
            home += 1
        if cell.value == "Health and beauty":
            home += 1



    wb.create_sheet('CategotySales')
    new_sheet = wb['CategotySales']
    data = (
        ("Category","Total Products Sold"),
        ("Fashion accessories",fashion),
        ("Food and beverages", food),
        ("Electronic accessories", electronic),
        ("Sports and travel", sport),
        ("Home and lifestyle", home)

    )
    for i in data:
        new_sheet.append(i)

    wb.save(exportFilename)


